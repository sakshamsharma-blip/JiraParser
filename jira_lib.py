"""
Core logic: parse Jira keys from sheets, fetch tickets, categorize, export.
Used by app.py (UI) and jira_changelog.py (CLI).
"""

from __future__ import annotations

import base64
import csv
import io
import json
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable

ROOT = Path(__file__).resolve().parent
ENV_PATH = ROOT / ".env"
CATEGORY_MAP_PATH = ROOT / "category_map.json"
OUT_DIR = ROOT / "output"
IMAGES_DIR = OUT_DIR / "images"

JIRA_KEY_RE = re.compile(r"\b([A-Z][A-Z0-9]+-\d+)\b")
IMAGE_MIME_PREFIX = "image/"

ProgressCb = Callable[[str], None]


def read_dotenv(path: Path | None = None) -> dict[str, str]:
    path = path or ENV_PATH
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        env[key.strip()] = value.strip().strip('"').strip("'")
    if "JIRA_BASE_URL" in env:
        env["JIRA_BASE_URL"] = env["JIRA_BASE_URL"].rstrip("/")
    return env


def sheet_csv_url(sheet_url: str) -> str:
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", sheet_url)
    if not m:
        raise ValueError("Not a Google Sheets URL. Paste a link like https://docs.google.com/spreadsheets/d/…")
    sheet_id = m.group(1)
    gid_match = re.search(r"[#&?]gid=(\d+)", sheet_url)
    gid = gid_match.group(1) if gid_match else "0"
    return (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}"
        f"/export?format=csv&gid={gid}"
    )


def http_get(url: str, headers: dict[str, str] | None = None, timeout: int = 60) -> bytes:
    req = urllib.request.Request(url, headers=headers or {})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read()
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")[:500]
        raise RuntimeError(f"HTTP {e.code}: {body}") from e


def _safe_filename(name: str) -> str:
    name = Path(name).name
    cleaned = re.sub(r"[^\w.\-()+ ]+", "_", name).strip()
    return cleaned or "image.bin"


def rows_from_csv_text(text: str) -> list[dict[str, str]]:
    reader = csv.DictReader(io.StringIO(text))
    rows = [dict(r) for r in reader]
    if not rows:
        raise ValueError("File/sheet has no data rows.")
    return rows


def rows_from_google_sheet(sheet_url: str) -> list[dict[str, str]]:
    raw = http_get(sheet_csv_url(sheet_url))
    return rows_from_csv_text(raw.decode("utf-8-sig", errors="replace"))


def rows_from_csv_bytes(data: bytes) -> list[dict[str, str]]:
    return rows_from_csv_text(data.decode("utf-8-sig", errors="replace"))


def rows_from_xlsx_bytes(data: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl is required for Excel files. Run: pip install -r requirements.txt") from e

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        headers = next(it)
    except StopIteration as e:
        raise ValueError("Excel sheet is empty.") from e
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(headers)]
    rows: list[dict[str, str]] = []
    for row in it:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        item: dict[str, str] = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else ""
            item[h] = "" if val is None else str(val)
        rows.append(item)
    if not rows:
        raise ValueError("Excel sheet has headers but no data rows.")
    return rows


def extract_keys_from_rows(rows: list[dict[str, str]]) -> list[str]:
    seen: set[str] = set()
    keys: list[str] = []
    for row in rows:
        blob = " ".join(str(v or "") for v in row.values())
        for key in JIRA_KEY_RE.findall(blob):
            if key not in seen:
                seen.add(key)
                keys.append(key)
    if not keys:
        raise ValueError("No Jira keys found (expected PROJ-123 or …/browse/PROJ-123).")
    return keys


def jira_auth_header(email: str, token: str) -> str:
    raw = f"{email}:{token}".encode("utf-8")
    return "Basic " + base64.b64encode(raw).decode("ascii")


def adf_to_text(node) -> str:
    if node is None:
        return ""
    if isinstance(node, str):
        return node
    if isinstance(node, list):
        return " ".join(adf_to_text(n) for n in node)
    if not isinstance(node, dict):
        return str(node)
    parts: list[str] = []
    if node.get("type") == "text":
        parts.append(node.get("text") or "")
    if "content" in node:
        parts.append(adf_to_text(node["content"]))
    if node.get("type") in {"paragraph", "heading", "bulletList", "orderedList", "listItem"}:
        parts.append("\n")
    return re.sub(r"[ \t]+\n", "\n", "".join(parts)).strip()


def fetch_issue(base_url: str, auth: str, key: str) -> dict:
    fields = ",".join(
        [
            "summary",
            "description",
            "issuetype",
            "status",
            "components",
            "labels",
            "priority",
            "created",
            "updated",
            "resolutiondate",
            "parent",
            "fixVersions",
            "attachment",
        ]
    )
    url = f"{base_url.rstrip('/')}/rest/api/3/issue/{urllib.parse.quote(key)}?fields={fields}"
    raw = http_get(url, headers={"Authorization": auth, "Accept": "application/json"})
    return json.loads(raw.decode("utf-8"))


def download_issue_images(issue: dict, auth: str) -> list[dict]:
    """Download image attachments for one issue into output/images/{KEY}/."""
    key = issue.get("key") or "UNKNOWN"
    attachments = (issue.get("fields") or {}).get("attachment") or []
    saved: list[dict] = []
    ticket_dir = IMAGES_DIR / key
    for att in attachments:
        mime = (att.get("mimeType") or "").lower()
        if not mime.startswith(IMAGE_MIME_PREFIX):
            continue
        content_url = att.get("content")
        filename = _safe_filename(att.get("filename") or f"{att.get('id', 'image')}.png")
        if not content_url:
            continue
        try:
            data = http_get(content_url, headers={"Authorization": auth, "Accept": "*/*"})
        except Exception:  # noqa: BLE001
            continue
        ticket_dir.mkdir(parents=True, exist_ok=True)
        dest = ticket_dir / filename
        # Avoid overwrite collisions
        if dest.exists():
            dest = ticket_dir / f"{att.get('id', 'x')}_{filename}"
        dest.write_bytes(data)
        rel = dest.relative_to(OUT_DIR).as_posix()
        saved.append(
            {
                "filename": dest.name,
                "rel_path": rel,
                "mime": mime,
                "id": str(att.get("id") or ""),
            }
        )
    return saved


def load_category_map(path: Path | None = None) -> dict:
    return json.loads((path or CATEGORY_MAP_PATH).read_text(encoding="utf-8"))


def categorize(issue: dict, cmap: dict) -> tuple[str, str]:
    fields = issue.get("fields") or {}
    summary = (fields.get("summary") or "").lower()
    description = adf_to_text(fields.get("description")).lower()
    blob = f"{summary} {description}"

    components = [c.get("name", "") for c in (fields.get("components") or [])]
    for comp in components:
        mapped = cmap.get("component_map", {}).get(comp.lower().strip())
        if mapped:
            return mapped, f"component:{comp}"
        for needle, stage in cmap.get("component_map", {}).items():
            if needle in comp.lower():
                return stage, f"component:{comp}"

    labels = [lab.lower() for lab in (fields.get("labels") or [])]
    for lab in labels:
        mapped = cmap.get("label_map", {}).get(lab)
        if mapped:
            return mapped, f"label:{lab}"

    for stage, keywords in cmap.get("keyword_map", {}).items():
        for kw in keywords:
            if kw.lower() in blob:
                return stage, f"keyword:{kw}"

    return "Needs review", "no-match"


def normalize_issue(issue: dict, cmap: dict, base_url: str) -> dict:
    fields = issue.get("fields") or {}
    key = issue.get("key")
    stage, reason = categorize(issue, cmap)
    parent = fields.get("parent") or {}
    description = adf_to_text(fields.get("description"))
    summary = fields.get("summary") or ""
    return {
        "key": key,
        "url": f"{base_url.rstrip('/')}/browse/{key}",
        "summary": summary,
        "description": description,
        "change_summary": _short_change_summary(summary, description),
        "type": ((fields.get("issuetype") or {}).get("name")) or "",
        "status": ((fields.get("status") or {}).get("name")) or "",
        "priority": ((fields.get("priority") or {}).get("name")) or "",
        "components": [c.get("name") for c in (fields.get("components") or [])],
        "labels": fields.get("labels") or [],
        "fixVersions": [v.get("name") for v in (fields.get("fixVersions") or [])],
        "parent_key": parent.get("key"),
        "parent_summary": (parent.get("fields") or {}).get("summary"),
        "created": fields.get("created"),
        "updated": fields.get("updated"),
        "resolved": fields.get("resolutiondate"),
        "category": stage,
        "category_reason": reason,
        "images": [],
    }


def _short_change_summary(summary: str, description: str, limit: int = 280) -> str:
    text = (description or "").strip() or (summary or "").strip()
    text = re.sub(r"\s+", " ", text)
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "…"


def fetch_and_categorize(
    keys: list[str],
    *,
    email: str,
    token: str,
    base_url: str,
    cmap: dict | None = None,
    progress: ProgressCb | None = None,
    delay_sec: float = 0.15,
    download_images: bool = True,
) -> tuple[list[dict], list[str]]:
    cmap = cmap or load_category_map()
    auth = jira_auth_header(email, token)
    base_url = base_url.rstrip("/")
    tickets: list[dict] = []
    failures: list[str] = []
    total = len(keys)
    for i, key in enumerate(keys, 1):
        if progress:
            progress(f"Fetching {i}/{total}: {key}")
        try:
            issue = fetch_issue(base_url, auth, key)
            ticket = normalize_issue(issue, cmap, base_url)
            if download_images:
                ticket["images"] = download_issue_images(issue, auth)
            tickets.append(ticket)
        except Exception as e:  # noqa: BLE001
            failures.append(f"{key}: {e}")
        time.sleep(delay_sec)
    return tickets, failures


def group_by_category(tickets: list[dict], cmap: dict | None = None) -> dict[str, list[dict]]:
    cmap = cmap or load_category_map()
    by_cat: dict[str, list[dict]] = defaultdict(list)
    for t in tickets:
        by_cat[t["category"]].append(t)
    stage_order = list(cmap.get("stages") or [])
    extras = sorted(c for c in by_cat if c not in stage_order)
    ordered = {c: by_cat[c] for c in stage_order if c in by_cat}
    for c in extras:
        ordered[c] = by_cat[c]
    return ordered


def build_markdown(tickets: list[dict], cmap: dict | None = None) -> str:
    cmap = cmap or load_category_map()
    by_cat = group_by_category(tickets, cmap)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    lines: list[str] = [
        "# Product changes from Jira",
        "",
        f"_Generated {now}._",
        "",
        f"**Total tickets:** {len(tickets)}",
        "",
        "## Category counts",
        "",
    ]
    for cat, items in by_cat.items():
        lines.append(f"- **{cat}:** {len(items)}")
    lines.append("")

    for cat, items in by_cat.items():
        lines.append(f"## {cat}")
        lines.append("")
        for t in items:
            lines.append(f"### [{t['key']}]({t['url']}) — {t['summary']}")
            lines.append("")
            meta = [
                f"**Type:** {t['type'] or '—'}",
                f"**Status:** {t['status'] or '—'}",
                f"**Categorized by:** `{t['category_reason']}`",
            ]
            if t["components"]:
                meta.append(f"**Components:** {', '.join(t['components'])}")
            if t["labels"]:
                meta.append(f"**Labels:** {', '.join(t['labels'])}")
            if t["fixVersions"]:
                meta.append(f"**Fix versions:** {', '.join(t['fixVersions'])}")
            lines.append(" · ".join(meta))
            lines.append("")
            if t.get("change_summary"):
                lines.append(f"**Change:** {t['change_summary']}")
                lines.append("")
            desc = (t.get("description") or "").strip()
            if desc and desc != t.get("change_summary"):
                if len(desc) > 1200:
                    desc = desc[:1200].rstrip() + "…"
                lines.append("**Description (from Jira):**")
                lines.append("")
                lines.append(desc)
                lines.append("")
            images = t.get("images") or []
            if images:
                lines.append("**Screenshots / images:**")
                lines.append("")
                for img in images:
                    lines.append(f"![{img['filename']}]({img['rel_path']})")
                    lines.append("")
            lines.append("---")
            lines.append("")
    return "\n".join(lines)


def build_csv(tickets: list[dict]) -> str:
    buf = io.StringIO()
    fields = [
        "key",
        "url",
        "category",
        "summary",
        "change_summary",
        "type",
        "status",
        "components",
        "labels",
        "fixVersions",
        "resolved",
        "category_reason",
    ]
    writer = csv.DictWriter(buf, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for t in tickets:
        row = dict(t)
        row["components"] = "; ".join(t.get("components") or [])
        row["labels"] = "; ".join(t.get("labels") or [])
        row["fixVersions"] = "; ".join(t.get("fixVersions") or [])
        writer.writerow(row)
    return buf.getvalue()


def build_pdf(tickets: list[dict], cmap: dict | None = None) -> bytes:
    """Build a PDF changelog with screenshots embedded (not just filenames)."""
    try:
        from fpdf import FPDF
        from fpdf.enums import XPos, YPos
    except ImportError as e:
        raise RuntimeError("fpdf2 is required for PDF export. Run: pip install -r requirements.txt") from e

    cmap = cmap or load_category_map()
    by_cat = group_by_category(tickets, cmap)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    class _PDF(FPDF):
        def footer(self) -> None:
            self.set_y(-12)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(120, 120, 120)
            self.cell(0, 8, f"Page {self.page_no()}/{{nb}}", align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    def write(pdf: "_PDF", text: str, *, size: int = 10, bold: bool = False, color=(0, 0, 0)) -> None:
        pdf.set_x(pdf.l_margin)
        style = "B" if bold else ""
        pdf.set_font("Helvetica", style, size)
        pdf.set_text_color(*color)
        pdf.multi_cell(0, max(5, size * 0.55), _pdf_safe(text), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf = _PDF(format="A4")
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.set_margins(14, 14, 14)
    pdf.add_page()

    write(pdf, "Product changes from Jira", size=18, bold=True)
    write(pdf, f"Generated {now}", size=10, color=(90, 90, 90))
    write(
        pdf,
        f"Total tickets: {len(tickets)}  |  Images: {count_images(tickets)}",
        size=10,
        color=(90, 90, 90),
    )
    pdf.ln(3)

    write(pdf, "Category counts", size=12, bold=True)
    for cat, items in by_cat.items():
        write(pdf, f"- {cat}: {len(items)}", size=10)
    pdf.ln(4)

    for cat, items in by_cat.items():
        pdf.set_x(pdf.l_margin)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_fill_color(235, 235, 235)
        pdf.set_text_color(0, 0, 0)
        pdf.multi_cell(0, 8, _pdf_safe(cat), fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(2)

        for t in items:
            write(pdf, f"{t.get('key') or ''} — {t.get('summary') or ''}", size=11, bold=True)

            meta_bits = [
                f"Type: {t.get('type') or '—'}",
                f"Status: {t.get('status') or '—'}",
            ]
            if t.get("components"):
                meta_bits.append("Components: " + ", ".join(t["components"]))
            write(pdf, "  |  ".join(meta_bits), size=9, color=(80, 80, 80))

            change = (t.get("change_summary") or "").strip()
            if change:
                write(pdf, "Change", size=10, bold=True)
                write(pdf, change, size=10)

            desc = (t.get("description") or "").strip()
            if desc and desc != change:
                if len(desc) > 1500:
                    desc = desc[:1500].rstrip() + "…"
                write(pdf, "Description", size=10, bold=True)
                write(pdf, desc, size=10)

            images = t.get("images") or []
            if images:
                write(pdf, "Screenshots", size=10, bold=True)
                usable_w = pdf.w - pdf.l_margin - pdf.r_margin
                for img in images:
                    img_path = OUT_DIR / img["rel_path"]
                    if not img_path.exists():
                        continue
                    if pdf.get_y() > pdf.h - 60:
                        pdf.add_page()
                    try:
                        pdf.set_x(pdf.l_margin)
                        pdf.image(str(img_path), w=min(170, usable_w))
                        pdf.ln(4)
                    except Exception:  # noqa: BLE001
                        write(pdf, f"(Could not embed {img.get('filename')})", size=9, color=(120, 120, 120))

            if t.get("url"):
                write(pdf, t["url"], size=8, color=(50, 90, 160))

            pdf.ln(2)
            y = pdf.get_y()
            pdf.set_draw_color(210, 210, 210)
            pdf.line(pdf.l_margin, y, pdf.w - pdf.r_margin, y)
            pdf.ln(5)

    return bytes(pdf.output())


def _pdf_safe(text: str) -> str:
    """FPDF core fonts are Latin-1; replace unsupported chars."""
    if not text:
        return ""
    return (
        text.replace("\u2014", "-")
        .replace("\u2013", "-")
        .replace("\u2018", "'")
        .replace("\u2019", "'")
        .replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("\u2026", "...")
        .encode("latin-1", errors="replace")
        .decode("latin-1")
    )


def save_outputs(tickets: list[dict], failures: list[str], cmap: dict | None = None) -> tuple[Path, Path, Path]:
    cmap = cmap or load_category_map()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    md_path = OUT_DIR / "changes-from-jira.md"
    json_path = OUT_DIR / "changes-from-jira.json"
    pdf_path = OUT_DIR / "changes-from-jira.pdf"
    md_path.write_text(build_markdown(tickets, cmap), encoding="utf-8")
    json_path.write_text(
        json.dumps(
            {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "ticket_count": len(tickets),
                "failures": failures,
                "tickets": tickets,
            },
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    pdf_path.write_bytes(build_pdf(tickets, cmap))
    return md_path, json_path, pdf_path



AI_REWRITE_SYSTEM = """You rewrite raw Jira change notes into a clear support/product changelog.

Rules:
- Group by the existing ## category headings.
- For each ticket, keep the key and title.
- Use short bullet points only: What changed, Who it affects / where in product (if known), Caveats (if any).
- Keep any markdown image links exactly as they appear (![...](...)). Put them under the ticket they belong to.
- Do not invent features that are not in the source.
- Output valid Markdown only. No preamble.
"""

AI_PROVIDERS = {
    "openai": {
        "label": "OpenAI",
        "default_model": "gpt-4o-mini",
        "models": ["gpt-4o-mini", "gpt-4o", "gpt-4.1-mini"],
        "default_base": "https://api.openai.com/v1",
        "env_key": "OPENAI_API_KEY",
        "env_model": "OPENAI_MODEL",
        "env_base": "OPENAI_API_BASE",
    },
    "gemini": {
        "label": "Google Gemini",
        # flash-lite usually has higher free-tier headroom than 2.0-flash
        "default_model": "gemini-2.0-flash-lite",
        "models": [
            "gemini-2.0-flash-lite",
            "gemini-2.0-flash",
            "gemini-2.5-flash",
            "gemini-1.5-flash",
        ],
        "default_base": "https://generativelanguage.googleapis.com/v1beta",
        "env_key": "GEMINI_API_KEY",
        "env_model": "GEMINI_MODEL",
        "env_base": "GEMINI_API_BASE",
    },
}


def _friendly_http_error(provider: str, code: int, body: str) -> str:
    lower = body.lower()
    if code == 429 or "resource_exhausted" in lower or "quota" in lower:
        retry = None
        m = re.search(r'"retryDelay"\s*:\s*"(\d+)s"', body)
        if m:
            retry = m.group(1)
        tip = (
            f"{provider} rate limit / free-tier quota hit. "
            "Wait a minute and try again, switch to a lighter model "
            "(e.g. gemini-2.0-flash-lite), or use OpenAI / a paid Gemini plan."
        )
        if retry:
            tip += f" Suggested wait: ~{retry}s."
        return tip
    if code in (401, 403):
        return f"{provider} auth failed (HTTP {code}). Check the API key."
    return f"{provider} API HTTP {code}: {body[:400]}"


def _openai_chat(api_key: str, system: str, user: str, model: str, base_url: str) -> str:
    url = base_url.rstrip("/") + "/chat/completions"
    payload = json.dumps(
        {
            "model": model,
            "messages": [
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            "temperature": 0.2,
        }
    ).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(_friendly_http_error("OpenAI", e.code, body)) from e
    try:
        return data["choices"][0]["message"]["content"].strip()
    except (KeyError, IndexError, TypeError) as e:
        raise RuntimeError(f"Unexpected OpenAI response: {data!r}") from e


def _gemini_generate(api_key: str, system: str, user: str, model: str, base_url: str) -> str:
    base = base_url.rstrip("/")
    url = (
        f"{base}/models/{urllib.parse.quote(model, safe='')}:generateContent"
        f"?key={urllib.parse.quote(api_key)}"
    )
    payload = json.dumps(
        {
            "systemInstruction": {"parts": [{"text": system}]},
            "contents": [{"role": "user", "parts": [{"text": user}]}],
            "generationConfig": {"temperature": 0.2},
        }
    ).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(_friendly_http_error("Gemini", e.code, body)) from e

    try:
        parts = data["candidates"][0]["content"]["parts"]
        texts = [p.get("text", "") for p in parts if isinstance(p, dict)]
        text = "".join(texts).strip()
        if not text:
            raise KeyError("empty")
        return text
    except (KeyError, IndexError, TypeError) as e:
        raise RuntimeError(f"Unexpected Gemini response: {data!r}") from e


def _ai_complete(
    provider: str,
    api_key: str,
    system: str,
    user: str,
    *,
    model: str | None = None,
    api_base: str | None = None,
    retries: int = 3,
) -> str:
    meta = AI_PROVIDERS.get(provider)
    if not meta:
        raise ValueError(f"Unsupported AI provider: {provider}. Use: {', '.join(AI_PROVIDERS)}")
    model = model or meta["default_model"]
    api_base = api_base or meta["default_base"]

    last_err: Exception | None = None
    for attempt in range(retries):
        try:
            if provider == "openai":
                return _openai_chat(api_key, system, user, model, api_base)
            if provider == "gemini":
                return _gemini_generate(api_key, system, user, model, api_base)
            raise ValueError(f"Unsupported AI provider: {provider}")
        except RuntimeError as e:
            last_err = e
            msg = str(e).lower()
            if attempt < retries - 1 and ("rate limit" in msg or "quota" in msg or "429" in msg):
                wait_m = re.search(r"~(\d+)s", str(e))
                wait = int(wait_m.group(1)) if wait_m else 20 * (attempt + 1)
                wait = min(max(wait, 5), 90)
                time.sleep(wait)
                continue
            raise
    assert last_err is not None
    raise last_err


def rewrite_structured_markdown(
    source_md: str,
    *,
    api_key: str,
    provider: str = "openai",
    model: str | None = None,
    api_base: str | None = None,
) -> str:
    """Turn raw changelog MD into structured bullet-point MD (OpenAI or Gemini)."""
    if not api_key.strip():
        raise ValueError("AI API key is required for rewrite.")

    provider = (provider or "openai").strip().lower()
    meta = AI_PROVIDERS.get(provider)
    if not meta:
        raise ValueError(f"Unsupported AI provider: {provider}")

    model = model or meta["default_model"]
    api_base = api_base or meta["default_base"]

    sections = re.split(r"(?=^## )", source_md, flags=re.MULTILINE)
    header = sections[0] if sections else ""
    body_sections = [s for s in sections[1:] if s.strip()]

    if not body_sections:
        chunks = [source_md]
    elif len(source_md) < 24000:
        chunks = [source_md]
    else:
        chunks = []
        buf = header
        for sec in body_sections:
            if len(buf) + len(sec) > 24000 and buf.strip():
                chunks.append(buf)
                buf = header + "\n" + sec
            else:
                buf += sec
        if buf.strip():
            chunks.append(buf)

    rewritten_parts: list[str] = []
    for i, chunk in enumerate(chunks, 1):
        user = (
            f"Rewrite this changelog section ({i}/{len(chunks)}) "
            "into structured bullet points:\n\n" + chunk
        )
        content = _ai_complete(
            provider,
            api_key,
            AI_REWRITE_SYSTEM,
            user,
            model=model,
            api_base=api_base,
        )
        rewritten_parts.append(content)
        if i < len(chunks):
            time.sleep(1.0)

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    out = [
        "# What changed (structured)",
        "",
        f"_AI rewrite ({meta['label']}, {model}) generated {now}._",
        "",
    ]
    out.extend(rewritten_parts)
    text = "\n\n".join(out).strip() + "\n"
    path = OUT_DIR / "changes-structured.md"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
    return text


def count_images(tickets: list[dict]) -> int:
    return sum(len(t.get("images") or []) for t in tickets)
